from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook,Workbook
import os
def log(tag, text):
	# Info tag
	if(tag == 'i'):
		print("[INFO] " + text)
	# Error tag
	elif(tag == 'e'):
		print("[ERROR] " + text)
	# Success tag
	elif(tag == 's'):
		print("[SUCCESS] " + text)

def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        sheet.append(_DATA)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = '시트이름'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 50
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 20
        sheet.append(_DATA)
        book.save(_FILENAME)

def get_bs_by_txt(_FILENAME):
    with open(_FILENAME,'r',encoding='utf8') as f:
        return BeautifulSoup(f.read(),'lxml')

cafeIdList = [585,586,392,418,590,591,592,242,126]
titleList = ['임신준비 질문방', '테스터,초음파 질문방', '임신 중 질문방', '난임,인공,시험관 질문방', '분만 질문방', '산후조리 질문방', '산후다이어트 질문방', '수유 질문방', '육아 질문방']
linkList = [
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=585&search.boardtype=L',
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=586&search.boardtype=L',
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=392&search.boardtype=L',
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=418&search.boardtype=L',
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=590&search.boardtype=L',
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=591&search.boardtype=L',
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=592&search.boardtype=L',
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=242&search.boardtype=L',
    'https://cafe.naver.com/imsanbu/ArticleList.nhn?search.clubid=10094499&search.menuid=126&search.boardtype=L'
    ]
urlFormat = 'https://cafe.naver.com/ArticleSearchList.nhn?search.clubid=10094499&search.searchdate={}{}&search.searchBy=0&search.query=&search.defaultValue=1&search.sortBy=date&userDisplay=50&search.media=0&search.option=0&search.menuid={}&search.page={}'
iframeUrl = 'https://cafe.naver.com/iframe_url=/imsanbu'

FILENAME = '{}_{}.xlsx'
HEADER = ['글번호', '작성일', '작성자', '제목', '내용', '답글1', '답글2', '답글3', '답글4', '답글5']
if __name__ =="__main__":
    bs4 = get_bs_by_txt('html.txt')
    id = bs4.find('a',id='linkUrl').get_text().strip().split('/')[-1]
    datetime = bs4.find('td',class_='m-tcol-c date').get_text().strip()
    author = bs4.find("div",class_='etc-box').find('td',class_='p-nick').a.get_text().strip()
    title = bs4.find('div',class_='tit-box').find('span',class_='b m-tcol-c').get_text().strip()
    contentDiv = bs4.find('div',id='tbody')
    try:
        contentDiv.find('div',class_='NHN_Writeform_Main').decompose()
    except:
        pass
    content = contentDiv.get_text().strip()
    commentList = ['','','','','']
    commentCnt = 0
    lis = bs4.find('ul',id='cmt_list').find_all('li',class_='')[:5]
    for li in lis:
        commentList[commentCnt] = li.find('span',class_='comm_body').get_text().strip()
        commentCnt+=1
    data= [id, datetime, author, title, content, commentList[0], commentList[1], commentList[2], commentList[3],commentList[4]]
    print(data)
    save_excel(FILENAME.format(datetime[:7]),data,HEADER)

